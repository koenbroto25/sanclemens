"""
Script untuk mengubah ribuan file txt "Data Umat" (umat.kasri.id) menjadi
1 file Excel rapi berisi DATA PRIBADI saja (1 baris per file/ID).
Data Keluarga (Kartu Keluarga) TIDAK diikutkan -- nanti bisa di-match
belakangan antar baris menggunakan kolom "No. Kartu Keluarga" yang sama.

Parsing dilakukan berdasarkan NAMA LABEL + posisi SECTION (Biodata /
Sakramen Baptis / Komuni / Penguatan / Perkawinan), bukan berdasarkan
indeks tetap. Ini supaya tahan terhadap variasi seperti baptisan di
Gereja Protestan yang punya 3 field tambahan (Gereja Protestan,
Tanggal Diterima, Penerima) dibanding baptisan Katolik biasa.

Cara pakai:
    python parse_umat.py <folder_input_txt> <output.xlsx>

Contoh:
    python parse_umat.py /mnt/user-data/uploads /mnt/user-data/outputs/data_umat.xlsx
"""

import sys
import re
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Kolom final Excel, urut sesuai tampilan biodata di sumber data.
# "(Sakramen)" membedakan field section Sakramen Baptis dari field
# biodata yang nama labelnya sama (Nama Baptis / Buku Baptis / Nomor Baptis).
COLUMN_HEADERS = [
    "ID", "Nama", "Paroki", "Wilayah", "Lingkungan", "Nama Baptis", "Agama",
    "Keuskupan Baptis", "Paroki Baptis", "Buku Baptis", "Nomor Baptis",
    "NIK", "No. Kartu Keluarga", "Tempat Lahir", "Tanggal Lahir",
    "Jenis Kelamin", "Golongan Darah", "Suku", "Hubungan Keluarga",
    "Status Perkawinan", "Tanggal Menikah", "Alamat", "Kota/Kab",
    "Kecamatan", "Kelurahan", "Status Rumah", "Handphone/Telepon",
    "Email", "Pendidikan Terakhir", "Fakultas/Jurusan", "Pekerjaan",
    "Profesi", "Keterampilan", "Kondisi Tubuh", "Status Ekonomi",
    "Status Aktivitas Sosial",
    "Nama Baptis (Sakramen)", "Buku Baptis (Sakramen)", "Nomor Baptis (Sakramen)",
    "Keuskupan Baptis/Penerima", "Paroki Baptis/Penerima", "Yang Membaptis",
    "Tanggal Baptis", "Baptis Gereja", "Wali Baptis",
    "Gereja Protestan", "Tanggal Diterima", "Penerima",  # field tambahan jika baptis Protestan
    "Notanda",
    "Buku Komuni Pertama", "Nomer Komuni Pertama", "Tanggal Komuni Pertama",
    "Keuskupan Komuni Pertama", "Paroki Komuni Pertama",
    "Buku Penguatan", "Nomer Penguatan", "Tanggal Penguatan",
    "Keuskupan Penguatan", "Paroki Penguatan", "Nama Penguatan",
    "Yang Memberi Penguatan",
    "Buku Perkawinan", "Nomer Buku Perkawinan", "Nama Pasangan",
    "Tanggal Perkawinan", "Jenis Perkawinan", "Keuskupan Perkawinan",
    "Paroki Perkawinan",
    "Update Terakhir", "Diupdate Oleh", "URL",
]

# Label biodata, urut sesuai kemunculan, sebelum kata "Sakramen".
# Section ini tidak punya label duplikat sehingga aman di-mapping 1:1.
BIODATA_LABELS = [
    "Nama", "Paroki", "Wilayah", "Lingkungan", "Nama Baptis", "Agama",
    "Keuskupan Baptis", "Paroki Baptis", "Buku Baptis", "Nomor Baptis",
    "NIK", "No. Kartu Keluarga", "Tempat Lahir", "Tanggal Lahir",
    "Jenis Kelamin", "Golongan Darah", "Suku", "Hubungan Keluarga",
    "Status Perkawinan", "Tanggal Menikah", "Alamat", "Kota/Kab",
    "Kecamatan", "Kelurahan", "Status Rumah", "Handphone/Telepon",
    "Email", "Pendidikan Terakhir", "Fakultas/Jurusan", "Pekerjaan",
    "Profesi", "Keterampilan", "Kondisi Tubuh", "Status Ekonomi",
    "Status Aktivitas Sosial",
]

# Label section Sakramen Baptis, urut sesuai kemunculan (field
# "Gereja Protestan"/"Tanggal Diterima"/"Penerima" bersifat opsional,
# ditangani lewat pencarian khusus, tidak lewat list fixed ini).
BAPTIS_LABELS_BASE = [
    "Nama Baptis", "Buku Baptis", "Nomor Baptis",
    "Keuskupan Baptis/Penerima", "Paroki Baptis/Penerima", "Yang Membaptis",
    "Tanggal Baptis", "Baptis Gereja", "Wali Baptis",
]
BAPTIS_OPTIONAL_LABELS = ["Gereja Protestan", "Tanggal Diterima", "Penerima"]

KOMUNI_LABELS = [
    "Buku Komuni Pertama", "Nomer Komuni Pertama", "Tanggal Komuni Pertama",
    "Keuskupan Komuni Pertama", "Paroki Komuni Pertama",
]
PENGUATAN_LABELS = [
    "Buku Penguatan", "Nomer Penguatan", "Tanggal Penguatan",
    "Keuskupan Penguatan", "Paroki Penguatan", "Nama Penguatan",
    "Yang Memberi Penguatan",
]
PERKAWINAN_LABELS = [
    "Buku Perkawinan", "Nomer Buku Perkawinan", "Nama Pasangan",
    "Tanggal Perkawinan", "Jenis Perkawinan", "Keuskupan Perkawinan",
    "Paroki Perkawinan",
]


def clean(val):
    """Bersihkan whitespace/tab/CR sisa; normalisasi nilai placeholder '-' jadi string kosong."""
    if val is None:
        return ""
    val = val.replace("\t", "").strip()
    if val == "-":
        return ""
    return val


def next_nonempty(lines, start_idx, n):
    """Cari index baris non-kosong pertama mulai dari start_idx (mengabaikan baris kosong)."""
    j = start_idx
    while j < n and lines[j].strip() == "":
        j += 1
    return j


def extract_labels_in_range(lines, start, end):
    """
    Ambil pasangan (label_tanpa_colon, value, index_label) untuk semua baris
    berakhiran ':' dalam rentang [start, end). Value = baris tepat sesudahnya.
    """
    out = []
    for i in range(start, end):
        l = lines[i].strip()
        if l.endswith(":"):
            label = l[:-1].strip()
            val = clean(lines[i + 1]) if i + 1 < end else ""
            out.append((label, val, i))
    return out


def parse_file(path):
    """Parse 1 file txt menjadi dict data pribadi (biodata + sakramen)."""
    with open(path, "rb") as f:
        raw = f.read()
    text = raw.decode("utf-8", errors="replace")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    lines = text.split("\n")
    n = len(lines)
    stripped = [l.strip() for l in lines]

    file_id = os.path.splitext(os.path.basename(path))[0]

    # --- ID & URL dari 2 baris pertama ---
    m = re.search(r"ID:\s*(\d+)", lines[0]) if n > 0 else None
    id_from_header = m.group(1) if m else file_id
    url = ""
    if n > 1 and stripped[1].startswith("URL:"):
        url = stripped[1].split("URL:", 1)[1].strip()

    # --- Update terakhir & oleh ---
    update_terakhir = ""
    diupdate_oleh = ""
    for i, l in enumerate(stripped):
        if l == "Update terakhir":
            j = next_nonempty(stripped, i + 1, n)
            update_terakhir = clean(lines[j]) if j < n else ""
            j2 = next_nonempty(stripped, j + 1, n)
            oleh_label = clean(lines[j2]) if j2 < n else ""
            if oleh_label == "oleh":
                j3 = next_nonempty(stripped, j2 + 1, n)
                diupdate_oleh = clean(lines[j3]) if j3 < n else ""
            break

    # --- Cari batas section: "Sakramen" dan "Keluarga" sebagai anchor ---
    try:
        sakramen_idx = stripped.index("Sakramen")
    except ValueError:
        sakramen_idx = n
    try:
        keluarga_idx = stripped.index("Keluarga")
    except ValueError:
        keluarga_idx = n

    data = {"ID": id_from_header}

    # --- Section Biodata: dari awal sampai sebelum "Sakramen" ---
    biodata_pairs = extract_labels_in_range(stripped, 0, sakramen_idx)
    # Hanya ambil label yang memang ada di BIODATA_LABELS, sesuai urutan kemunculan pertama
    biodata_map = {}
    for label, val, idx in biodata_pairs:
        if label in BIODATA_LABELS and label not in biodata_map:
            biodata_map[label] = val
    for label in BIODATA_LABELS:
        data[label] = biodata_map.get(label, "")

    # --- Section Sakramen: dari "Sakramen" sampai sebelum "Keluarga" ---
    sakramen_pairs = extract_labels_in_range(stripped, sakramen_idx, keluarga_idx)
    sakramen_labels_seen = [p[0] for p in sakramen_pairs]

    # cari posisi anchor sub-section dalam stripped (kata tunggal: Baptis/Komuni/Penguatan/Perkawinan)
    def find_subsection(name, after_idx):
        for i in range(after_idx, keluarga_idx):
            if stripped[i] == name:
                return i
        return None

    baptis_anchor = find_subsection("Baptis", sakramen_idx)
    komuni_anchor = find_subsection("Komuni", baptis_anchor or sakramen_idx)
    penguatan_anchor = find_subsection("Penguatan", komuni_anchor or sakramen_idx)
    perkawinan_anchor = find_subsection("Perkawinan", penguatan_anchor or sakramen_idx)

    # --- Sub-section Baptis: dari baptis_anchor sampai komuni_anchor ---
    baptis_end = komuni_anchor if komuni_anchor else keluarga_idx
    baptis_pairs = extract_labels_in_range(stripped, baptis_anchor or sakramen_idx, baptis_end)
    baptis_map = {}
    for label, val, idx in baptis_pairs:
        # hindari menimpa nilai pertama jika label sama muncul lagi (seharusnya tidak terjadi di section ini)
        if label not in baptis_map:
            baptis_map[label] = val

    data["Nama Baptis (Sakramen)"] = baptis_map.get("Nama Baptis", "")
    data["Buku Baptis (Sakramen)"] = baptis_map.get("Buku Baptis", "")
    data["Nomor Baptis (Sakramen)"] = baptis_map.get("Nomor Baptis", "")
    data["Keuskupan Baptis/Penerima"] = baptis_map.get("Keuskupan Baptis/Penerima", "")
    data["Paroki Baptis/Penerima"] = baptis_map.get("Paroki Baptis/Penerima", "")
    data["Yang Membaptis"] = baptis_map.get("Yang Membaptis", "")
    data["Tanggal Baptis"] = baptis_map.get("Tanggal Baptis", "")
    data["Baptis Gereja"] = baptis_map.get("Baptis Gereja", "")
    data["Wali Baptis"] = baptis_map.get("Wali Baptis", "")
    # Field opsional (hanya ada jika baptis di Gereja Protestan)
    data["Gereja Protestan"] = baptis_map.get("Gereja Protestan", "")
    data["Tanggal Diterima"] = baptis_map.get("Tanggal Diterima", "")
    data["Penerima"] = baptis_map.get("Penerima", "")
    data["Notanda"] = baptis_map.get("Notanda", "")

    # --- Sub-section Komuni ---
    komuni_end = penguatan_anchor if penguatan_anchor else keluarga_idx
    komuni_pairs = extract_labels_in_range(stripped, komuni_anchor or sakramen_idx, komuni_end)
    komuni_map = {label: val for label, val, idx in komuni_pairs if label not in {}}
    for label in KOMUNI_LABELS:
        data[label] = komuni_map.get(label, "")

    # --- Sub-section Penguatan ---
    penguatan_end = perkawinan_anchor if perkawinan_anchor else keluarga_idx
    penguatan_pairs = extract_labels_in_range(stripped, penguatan_anchor or sakramen_idx, penguatan_end)
    penguatan_map = {label: val for label, val, idx in penguatan_pairs}
    for label in PENGUATAN_LABELS:
        data[label] = penguatan_map.get(label, "")

    # --- Sub-section Perkawinan ---
    perkawinan_pairs = extract_labels_in_range(stripped, perkawinan_anchor or sakramen_idx, keluarga_idx)
    perkawinan_map = {label: val for label, val, idx in perkawinan_pairs}
    for label in PERKAWINAN_LABELS:
        data[label] = perkawinan_map.get(label, "")

    data["Update Terakhir"] = update_terakhir
    data["Diupdate Oleh"] = diupdate_oleh
    data["URL"] = url

    # --- Validasi: cek apakah semua label biodata wajib ditemukan ---
    missing_biodata = [l for l in BIODATA_LABELS if l not in biodata_map]
    extra_sakramen_labels = set(sakramen_labels_seen) - set(
        BAPTIS_LABELS_BASE + BAPTIS_OPTIONAL_LABELS + ["Notanda"]
        + KOMUNI_LABELS + PENGUATAN_LABELS + PERKAWINAN_LABELS
    )

    warning = None
    problems = []
    if missing_biodata:
        problems.append(f"label biodata hilang: {missing_biodata}")
    if extra_sakramen_labels:
        problems.append(f"label sakramen tak dikenal: {sorted(extra_sakramen_labels)}")
    if sakramen_idx == n:
        problems.append("section 'Sakramen' tidak ditemukan")
    if problems:
        warning = "; ".join(problems)

    return data, warning


def build_excel(input_folder, output_path):
    files = sorted(
        glob.glob(os.path.join(input_folder, "*.txt")),
        key=lambda p: (len(os.path.basename(p)), os.path.basename(p))
    )
    if not files:
        print(f"Tidak ada file .txt ditemukan di: {input_folder}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Biodata"

    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    header_fill = PatternFill("solid", start_color="4472C4")
    normal_font = Font(name="Arial", size=10)

    ws.append(COLUMN_HEADERS)
    for c in range(1, len(COLUMN_HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    # Mapping header Excel -> key di dalam dict `data`
    header_to_key = {h: h for h in COLUMN_HEADERS}

    error_log = []
    warning_log = []

    for path in files:
        try:
            data, warning = parse_file(path)
        except Exception as e:
            error_log.append(f"{os.path.basename(path)}: {e}")
            continue

        if warning:
            warning_log.append(f"{os.path.basename(path)}: {warning}")

        row = [data.get(header_to_key.get(h, h), "") for h in COLUMN_HEADERS]
        ws.append(row)

    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)
    for r in ws.iter_rows(min_row=2):
        for c in r:
            c.font = normal_font

    wb.save(output_path)

    print(f"Selesai. Total file diproses : {len(files)}")
    print(f"Total baris data              : {ws.max_row - 1}")
    if warning_log:
        print(f"\nFile dengan kejanggalan struktur ({len(warning_log)}) -- PERLU DICEK MANUAL:")
        for w in warning_log[:30]:
            print(" -", w)
        if len(warning_log) > 30:
            print(f" ... dan {len(warning_log) - 30} lainnya")
    if error_log:
        print(f"\nFile gagal diproses total error ({len(error_log)}):")
        for e in error_log[:30]:
            print(" -", e)
        if len(error_log) > 30:
            print(f" ... dan {len(error_log) - 30} lainnya")
    print(f"\nFile output: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Cara pakai: python parse_umat.py <folder_input_txt> <output.xlsx>")
        sys.exit(1)
    build_excel(sys.argv[1], sys.argv[2])
