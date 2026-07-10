"""
================================================================================
 SKRIP GEOCODING & KLASIFIKASI LINGKUNGAN - PAROKI ST. KLEMENS SEPINGGAN
================================================================================

FUNGSI SKRIP INI:
1. Membaca file Excel data umat (kolom: ID, Nama, Alamat, Kota/Kab, Kecamatan, Kelurahan)
2. Mengubah tiap alamat menjadi koordinat (latitude, longitude) lewat Google Geocoding API
3. Mengecek koordinat tersebut masuk ke poligon lingkungan mana
   (berdasarkan file lingkungan_sepinggan.geojson hasil konversi dari KML paroki)
4. Menyimpan hasil ke file Excel baru dengan kolom tambahan:
   Latitude, Longitude, Lingkungan, Status_Geocode

CARA PAKAI:
1. Pastikan sudah punya API KEY Google Geocoding (lihat panduan di README.md)
2. Install dependency:
       pip install openpyxl shapely requests
3. Isi API_KEY di bawah ini
4. Jalankan:
       python geocode_klasifikasi_lingkungan.py

CATATAN PENTING:
- Skrip ini punya fitur CACHE & RESUME: hasil geocoding disimpan bertahap ke
  file cache (geocode_cache.json) supaya kalau skrip terhenti di tengah jalan
  (mati listrik, koneksi putus, dll), bisa dilanjut tanpa mengulang dari awal
  dan tanpa boros kuota API.
- Ada jeda antar-request kecil untuk menghindari rate limit.
- Alamat yang gagal di-geocode atau berada di luar semua poligon lingkungan
  akan ditandai di kolom Status_Geocode supaya bisa dicek manual.
================================================================================
"""

import json
import time
import re
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill
from shapely.geometry import shape, Point

# ============================== KONFIGURASI ==================================

API_KEY = "ISI_API_KEY_GOOGLE_ANDA_DI_SINI"

FILE_INPUT = "data_umat-alamat.xlsx"          # file data umat
SHEET_NAME = "Biodata"                         # nama sheet di file excel
FILE_GEOJSON = "lingkungan_sepinggan.geojson"  # batas wilayah lingkungan
FILE_OUTPUT = "data_umat-alamat_HASIL.xlsx"    # file hasil akhir
FILE_CACHE = "geocode_cache.json"              # cache hasil geocoding (auto-dibuat)

# Konteks tambahan supaya geocoding lebih akurat (karena banyak alamat hanya
# nama jalan tanpa konteks kota). Sesuaikan jika perlu.
KONTEKS_WILAYAH = "Kalimantan Timur, Indonesia"

JEDA_ANTAR_REQUEST = 0.05   # detik, jeda antar panggilan API (hindari rate limit)

# ===============================================================================


def muat_poligon_lingkungan(path_geojson):
    """Membaca file GeoJSON dan mengembalikan list of (nama_lingkungan, shapely_polygon)."""
    with open(path_geojson, "r", encoding="utf-8") as f:
        geo = json.load(f)
    daftar = []
    for feat in geo["features"]:
        nama = feat["properties"]["lingkungan"]
        poly = shape(feat["geometry"])
        daftar.append((nama, poly))
    return daftar


def cari_lingkungan(lat, lon, daftar_poligon):
    """Mengecek titik (lat, lon) ada di poligon lingkungan mana. Return nama atau None."""
    titik = Point(lon, lat)  # shapely pakai urutan (x=lon, y=lat)
    for nama, poly in daftar_poligon:
        if poly.contains(titik):
            return nama
    return None


def bersihkan_alamat(alamat, kelurahan, kecamatan, kota):
    """Menggabungkan kolom alamat menjadi satu string query yang lebih lengkap untuk geocoding."""
    bagian = [str(x).strip() for x in [alamat, kelurahan, kecamatan, kota] if x and str(x).strip()]
    query = ", ".join(bagian) + ", " + KONTEKS_WILAYAH
    # rapikan spasi ganda
    query = re.sub(r"\s+", " ", query).strip()
    return query


def geocode_google(alamat_lengkap, api_key):
    """
    Memanggil Google Geocoding API.
    Return: (lat, lon, status, alamat_hasil_google) atau (None, None, status, None) jika gagal.
    """
    url = "https://maps.googleapis.com/maps/api/geocode/json"
    params = {"address": alamat_lengkap, "key": api_key, "region": "id"}
    try:
        resp = requests.get(url, params=params, timeout=10)
        data = resp.json()
    except Exception as e:
        return None, None, f"ERROR_KONEKSI: {e}", None

    status = data.get("status")
    if status == "OK":
        hasil = data["results"][0]
        lat = hasil["geometry"]["location"]["lat"]
        lon = hasil["geometry"]["location"]["lng"]
        alamat_formatted = hasil.get("formatted_address")
        return lat, lon, "OK", alamat_formatted
    else:
        # status umum: ZERO_RESULTS, OVER_QUERY_LIMIT, REQUEST_DENIED, INVALID_REQUEST
        return None, None, status, None


def muat_cache(path_cache):
    try:
        with open(path_cache, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return {}


def simpan_cache(cache, path_cache):
    with open(path_cache, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=0)


def main():
    if API_KEY == "ISI_API_KEY_GOOGLE_ANDA_DI_SINI" or not API_KEY:
        print("!! API_KEY belum diisi. Buka skrip ini, isi variabel API_KEY di bagian atas.")
        return

    print("Memuat poligon lingkungan...")
    daftar_poligon = muat_poligon_lingkungan(FILE_GEOJSON)
    print(f"  -> {len(daftar_poligon)} poligon lingkungan dimuat.")

    print("Membuka file data umat...")
    wb = openpyxl.load_workbook(FILE_INPUT)
    ws = wb[SHEET_NAME]

    header = [cell.value for cell in ws[1]]
    col = {nama: idx + 1 for idx, nama in enumerate(header)}  # nama_kolom -> nomor kolom (1-based)

    kolom_wajib = ["Alamat", "Kota/Kab", "Kecamatan", "Kelurahan"]
    for k in kolom_wajib:
        if k not in col:
            print(f"!! Kolom '{k}' tidak ditemukan di sheet. Cek nama kolom di file Excel.")
            return

    # tambah kolom baru jika belum ada
    kolom_baru = ["Latitude", "Longitude", "Lingkungan", "Status_Geocode", "Alamat_Hasil_Google"]
    next_col = ws.max_column + 1
    for nama_kolom in kolom_baru:
        if nama_kolom not in col:
            ws.cell(row=1, column=next_col, value=nama_kolom).font = Font(bold=True)
            col[nama_kolom] = next_col
            next_col += 1

    cache = muat_cache(FILE_CACHE)

    total = ws.max_row - 1
    print(f"Total data umat: {total}")
    print("Mulai proses geocoding (bisa berhenti & dilanjut kapan saja, progres tersimpan)...\n")

    jumlah_ok = 0
    jumlah_gagal = 0
    jumlah_diluar_wilayah = 0

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        id_umat = row[col["ID"] - 1].value if "ID" in col else i
        alamat = row[col["Alamat"] - 1].value
        kota = row[col["Kota/Kab"] - 1].value
        kecamatan = row[col["Kecamatan"] - 1].value
        kelurahan = row[col["Kelurahan"] - 1].value

        # lewati baris yang sudah pernah berhasil diproses sebelumnya (skip resume)
        status_existing = ws.cell(row=i, column=col["Status_Geocode"]).value
        if status_existing == "OK":
            continue

        query = bersihkan_alamat(alamat, kelurahan, kecamatan, kota)

        if query in cache:
            lat, lon, status, alamat_formatted = cache[query]
        else:
            lat, lon, status, alamat_formatted = geocode_google(query, API_KEY)
            cache[query] = (lat, lon, status, alamat_formatted)
            simpan_cache(cache, FILE_CACHE)
            time.sleep(JEDA_ANTAR_REQUEST)

        ws.cell(row=i, column=col["Latitude"], value=lat)
        ws.cell(row=i, column=col["Longitude"], value=lon)
        ws.cell(row=i, column=col["Alamat_Hasil_Google"], value=alamat_formatted)

        if status == "OK" and lat is not None:
            nama_lingkungan = cari_lingkungan(lat, lon, daftar_poligon)
            if nama_lingkungan:
                ws.cell(row=i, column=col["Lingkungan"], value=nama_lingkungan)
                ws.cell(row=i, column=col["Status_Geocode"], value="OK")
                jumlah_ok += 1
            else:
                ws.cell(row=i, column=col["Lingkungan"], value="DI LUAR BATAS PAROKI")
                ws.cell(row=i, column=col["Status_Geocode"], value="OK_TAPI_DILUAR_POLIGON")
                jumlah_diluar_wilayah += 1
        else:
            ws.cell(row=i, column=col["Status_Geocode"], value=f"GAGAL: {status}")
            jumlah_gagal += 1

        if (i - 1) % 50 == 0:
            print(f"  ... {i - 1}/{total} diproses "
                  f"(OK: {jumlah_ok}, di luar wilayah: {jumlah_diluar_wilayah}, gagal: {jumlah_gagal})")
            wb.save(FILE_OUTPUT)  # simpan progres berkala

    # pewarnaan sederhana untuk baris yang perlu dicek manual
    isi_kuning = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
    col_status = col["Status_Geocode"]
    for i in range(2, ws.max_row + 1):
        status_val = ws.cell(row=i, column=col_status).value
        if status_val and status_val != "OK":
            for c in range(1, ws.max_column + 1):
                ws.cell(row=i, column=c).fill = isi_kuning

    wb.save(FILE_OUTPUT)

    print("\n=== SELESAI ===")
    print(f"Berhasil & masuk lingkungan : {jumlah_ok}")
    print(f"Berhasil tapi di luar batas paroki (mis. luar Balikpapan) : {jumlah_diluar_wilayah}")
    print(f"Gagal di-geocode (perlu dicek manual) : {jumlah_gagal}")
    print(f"File hasil disimpan di: {FILE_OUTPUT}")
    print("Baris berwarna kuning di file hasil = perlu dicek/dilengkapi manual.")


if __name__ == "__main__":
    main()
