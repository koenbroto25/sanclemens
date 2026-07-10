"""
================================================================================
 SKRIP KLASIFIKASI LINGKUNGAN v2 - PAROKI ST. KLEMENS SEPINGGAN
 (Sistem 3 Lapis: Kata Kunci Wilayah -> Geocoding+Poligon -> Flag Manual)
================================================================================

CARA KERJA:

  LAPIS 1 - KATA KUNCI WILAYAH (gratis, instan, tanpa API)
      Alamat umat dicocokkan dengan "kamus wilayah" (kamus_wilayah.json) yang
      disusun dari deskripsi wilayah resmi tiap Ketua Lingkungan (nama jalan,
      perumahan, dan patokan). Kalau cocok dengan TEPAT SATU lingkungan -> langsung
      diklasifikasikan, tidak perlu panggil API sama sekali.

  LAPIS 2 - GEOCODING + POLIGON (untuk yang belum ketemu di Lapis 1, & cross-check)
      Alamat yang tidak cocok kata kunci (atau cocok ke >1 lingkungan / ambigu)
      diproses lewat Google Geocoding API lalu dicek masuk poligon lingkungan
      mana (dari lingkungan_sepinggan.geojson).

  LAPIS 3 - FLAG MANUAL DENGAN KONTAK KETUA LINGKUNGAN
      Kalau Lapis 1 & 2 tidak sepakat, atau dua-duanya gagal/ambigu, baris
      ditandai untuk dicek manual DAN otomatis dilengkapi nama+kontak ketua
      lingkungan kandidat supaya tim pendataan tinggal konfirmasi langsung.

KOLOM HASIL TAMBAHAN:
  - Lingkungan_KataKunci   : hasil Lapis 1 (atau kosong/"AMBIGU:..." )
  - Lingkungan_Geocoding   : hasil Lapis 2 (lewat koordinat & poligon)
  - Lingkungan_Final       : keputusan akhir gabungan
  - Tingkat_Keyakinan      : TINGGI (1&2 sepakat) / SEDANG (hanya 1 metode) / PERLU_CEK_MANUAL
  - Kontak_Untuk_Konfirmasi: nama & no. HP ketua lingkungan kandidat (utk kasus PERLU_CEK_MANUAL)
  - Latitude / Longitude   : hasil geocoding
  - Status_Geocode         : status mentah dari Google API

CARA PAKAI: sama seperti versi sebelumnya, lihat README.md
================================================================================
"""

import json
import time
import re
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill
from shapely.geometry import shape, Point

# ============================== KONFIGURASI ==================================

API_KEY = "ISI_API_KEY_GOOGLE_ANDA_DI_SINI"

FILE_INPUT = "data_umat-alamat.xlsx"
SHEET_NAME = "Biodata"
FILE_GEOJSON = "lingkungan_sepinggan.geojson"
FILE_KAMUS = "kamus_wilayah.json"
FILE_OUTPUT = "data_umat-alamat_HASIL.xlsx"
FILE_CACHE = "geocode_cache.json"

KONTEKS_WILAYAH = "Kalimantan Timur, Indonesia"
JEDA_ANTAR_REQUEST = 0.05

# Kalau True: alamat yang sudah dapat hasil TINGGI dari Lapis 1 akan TETAP
# digeocode juga untuk cross-check (lebih akurat tapi lebih lama & lebih banyak
# kuota API). Kalau False: alamat yang sudah pasti dari Lapis 1 langsung dipakai
# tanpa geocoding (lebih cepat & hemat kuota, direkomendasikan untuk data besar).
CROSS_CHECK_SEMUA = False

# ===============================================================================


def normalisasi(teks):
    """Menyamakan variasi penulisan: huruf kecil, hilangkan tanda baca & kata umum."""
    if not teks:
        return ""
    t = str(teks).lower()
    t = re.sub(r"[.,/\-]", " ", t)
    # samakan singkatan umum
    ganti = {
        r"\bjl\b": "jalan", r"\bgg\b": "gang", r"\bperum\b": "perumahan",
        r"\bno\b": "", r"\brt\b": "rt", r"\bgn\b": "gunung",
    }
    for pola, ganti_dengan in ganti.items():
        t = re.sub(pola, ganti_dengan, t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def muat_kamus(path):
    with open(path, "r", encoding="utf-8") as f:
        kamus = json.load(f)
    # pra-normalisasi keyword supaya pencocokan cepat
    for entri in kamus:
        entri["_keywords_norm"] = [normalisasi(k) for k in entri["keywords"]]
    return kamus


def cocokkan_kata_kunci(alamat_lengkap, kamus):
    """
    Mengembalikan list lingkungan yang kata kuncinya match di alamat.
    Bisa lebih dari satu (ambigu) atau kosong (tidak ketemu).
    """
    teks = normalisasi(alamat_lengkap)
    hasil = []
    for entri in kamus:
        for kw_norm, kw_asli in zip(entri["_keywords_norm"], entri["keywords"]):
            if len(kw_norm) < 3:
                continue
            # pencocokan kata/frasa utuh (word boundary) supaya "AD" tidak match sembarangan
            pola = r"\b" + re.escape(kw_norm) + r"\b"
            if re.search(pola, teks):
                hasil.append((entri["lingkungan"], entri["nama_kml"], entri["ketua"],
                               entri["kontak"], kw_asli))
                break  # cukup 1 keyword match per lingkungan, lanjut entri berikutnya
    return hasil


def muat_poligon_lingkungan(path_geojson):
    with open(path_geojson, "r", encoding="utf-8") as f:
        geo = json.load(f)
    return [(feat["properties"]["lingkungan"], shape(feat["geometry"])) for feat in geo["features"]]


def cari_lingkungan_poligon(lat, lon, daftar_poligon):
    titik = Point(lon, lat)
    for nama, poly in daftar_poligon:
        if poly.contains(titik):
            return nama
    return None


def bersihkan_alamat(alamat, kelurahan, kecamatan, kota):
    bagian = [str(x).strip() for x in [alamat, kelurahan, kecamatan, kota] if x and str(x).strip()]
    query = ", ".join(bagian) + ", " + KONTEKS_WILAYAH
    return re.sub(r"\s+", " ", query).strip()


def geocode_google(alamat_lengkap, api_key):
    url = "https://maps.googleapis.com/maps/api/geocode/json"
    params = {"address": alamat_lengkap, "key": api_key, "region": "id"}
    try:
        resp = requests.get(url, params=params, timeout=10)
        data = resp.json()
    except Exception as e:
        return None, None, f"ERROR_KONEKSI: {e}", None
    status = data.get("status")
    if status == "OK":
        hasil = data["results"][0]
        lat = hasil["geometry"]["location"]["lat"]
        lon = hasil["geometry"]["location"]["lng"]
        return lat, lon, "OK", hasil.get("formatted_address")
    return None, None, status, None


def muat_cache(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return {}


def simpan_cache(cache, path):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False)


def main():
    if API_KEY == "ISI_API_KEY_GOOGLE_ANDA_DI_SINI" or not API_KEY:
        print("!! API_KEY belum diisi (perlu untuk Lapis 2 - geocoding). "
              "Lapis 1 (kata kunci) tetap akan berjalan, tapi alamat yang tidak "
              "ketemu kata kuncinya tidak akan diproses lebih lanjut.")

    print("Memuat kamus wilayah (kata kunci per lingkungan)...")
    kamus = muat_kamus(FILE_KAMUS)
    print(f"  -> {len(kamus)} lingkungan, {sum(len(k['keywords']) for k in kamus)} kata kunci total.")

    print("Memuat poligon lingkungan...")
    daftar_poligon = muat_poligon_lingkungan(FILE_GEOJSON)
    print(f"  -> {len(daftar_poligon)} poligon dimuat.")

    print("Membuka file data umat...")
    wb = openpyxl.load_workbook(FILE_INPUT)
    ws = wb[SHEET_NAME]
    header = [cell.value for cell in ws[1]]
    col = {nama: idx + 1 for idx, nama in enumerate(header)}

    for k in ["Alamat", "Kota/Kab", "Kecamatan", "Kelurahan"]:
        if k not in col:
            print(f"!! Kolom '{k}' tidak ditemukan.")
            return

    kolom_baru = ["Lingkungan_KataKunci", "Lingkungan_Geocoding", "Lingkungan_Final",
                  "Tingkat_Keyakinan", "Kontak_Untuk_Konfirmasi",
                  "Latitude", "Longitude", "Status_Geocode"]
    next_col = ws.max_column + 1
    for nama_kolom in kolom_baru:
        if nama_kolom not in col:
            ws.cell(row=1, column=next_col, value=nama_kolom).font = Font(bold=True)
            col[nama_kolom] = next_col
            next_col += 1

    cache = muat_cache(FILE_CACHE)
    total = ws.max_row - 1
    print(f"Total data umat: {total}\n")

    hitung = {"katakunci_pasti": 0, "geocoding_dipakai": 0, "sepakat": 0,
              "ambigu_atau_gagal": 0}

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if ws.cell(row=i, column=col["Tingkat_Keyakinan"]).value == "TINGGI":
            continue  # sudah selesai dari proses sebelumnya (fitur resume)

        alamat = row[col["Alamat"] - 1].value
        kota = row[col["Kota/Kab"] - 1].value
        kecamatan = row[col["Kecamatan"] - 1].value
        kelurahan = row[col["Kelurahan"] - 1].value
        alamat_gabungan = f"{alamat} {kelurahan} {kecamatan} {kota}"

        # ---------- LAPIS 1: kata kunci ----------
        cocok = cocokkan_kata_kunci(alamat_gabungan, kamus)
        lingkungan_unik = list({c[0] for c in cocok})

        hasil_katakunci = None
        if len(lingkungan_unik) == 1:
            hasil_katakunci = lingkungan_unik[0]
            ws.cell(row=i, column=col["Lingkungan_KataKunci"], value=hasil_katakunci)
        elif len(lingkungan_unik) > 1:
            ws.cell(row=i, column=col["Lingkungan_KataKunci"],
                    value="AMBIGU: " + " / ".join(lingkungan_unik))
        # kalau kosong, biarkan kosong

        perlu_geocode = (hasil_katakunci is None or len(lingkungan_unik) != 1
                          or CROSS_CHECK_SEMUA)

        hasil_geocoding = None
        if perlu_geocode and API_KEY != "ISI_API_KEY_GOOGLE_ANDA_DI_SINI" and API_KEY:
            query = bersihkan_alamat(alamat, kelurahan, kecamatan, kota)
            if query in cache:
                lat, lon, status, _ = cache[query]
            else:
                lat, lon, status, alamat_fmt = geocode_google(query, API_KEY)
                cache[query] = (lat, lon, status, alamat_fmt)
                simpan_cache(cache, FILE_CACHE)
                time.sleep(JEDA_ANTAR_REQUEST)

            ws.cell(row=i, column=col["Latitude"], value=lat)
            ws.cell(row=i, column=col["Longitude"], value=lon)
            ws.cell(row=i, column=col["Status_Geocode"], value=status)
            hitung["geocoding_dipakai"] += 1

            if status == "OK" and lat is not None:
                nama_kml = cari_lingkungan_poligon(lat, lon, daftar_poligon)
                if nama_kml:
                    # balikkan nama_kml -> nama resmi SK lewat kamus
                    cocokan = next((k["lingkungan"] for k in kamus if k["nama_kml"] == nama_kml), nama_kml)
                    hasil_geocoding = cocokan
                    ws.cell(row=i, column=col["Lingkungan_Geocoding"], value=cocokan)
                else:
                    ws.cell(row=i, column=col["Lingkungan_Geocoding"], value="DI LUAR BATAS PAROKI")

        # ---------- KEPUTUSAN FINAL ----------
        if hasil_katakunci and hasil_geocoding and hasil_katakunci == hasil_geocoding:
            ws.cell(row=i, column=col["Lingkungan_Final"], value=hasil_katakunci)
            ws.cell(row=i, column=col["Tingkat_Keyakinan"], value="TINGGI")
            hitung["katakunci_pasti"] += 1
            hitung["sepakat"] += 1
        elif hasil_katakunci and not perlu_geocode:
            # hanya lapis 1, tidak digeocode (mode hemat kuota)
            ws.cell(row=i, column=col["Lingkungan_Final"], value=hasil_katakunci)
            ws.cell(row=i, column=col["Tingkat_Keyakinan"], value="TINGGI")
            hitung["katakunci_pasti"] += 1
        elif hasil_katakunci and not hasil_geocoding:
            ws.cell(row=i, column=col["Lingkungan_Final"], value=hasil_katakunci)
            ws.cell(row=i, column=col["Tingkat_Keyakinan"], value="SEDANG (hanya kata kunci)")
        elif hasil_geocoding and not hasil_katakunci:
            ws.cell(row=i, column=col["Lingkungan_Final"], value=hasil_geocoding)
            ws.cell(row=i, column=col["Tingkat_Keyakinan"], value="SEDANG (hanya geocoding)")
        else:
            # ambigu / tidak ketemu sama sekali / dua metode beda hasil
            ws.cell(row=i, column=col["Tingkat_Keyakinan"], value="PERLU_CEK_MANUAL")
            kandidat = lingkungan_unik if lingkungan_unik else (
                [hasil_geocoding] if hasil_geocoding else [])
            kontak_info = []
            for nama_lk in kandidat:
                entri = next((k for k in kamus if k["lingkungan"] == nama_lk), None)
                if entri:
                    kontak_info.append(f"{nama_lk}: {entri['ketua']} ({entri['kontak']})")
            ws.cell(row=i, column=col["Kontak_Untuk_Konfirmasi"], value=" | ".join(kontak_info))
            hitung["ambigu_atau_gagal"] += 1

        if (i - 1) % 100 == 0:
            print(f"  ... {i - 1}/{total} diproses")
            wb.save(FILE_OUTPUT)

    # pewarnaan baris yang perlu dicek manual
    isi_kuning = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
    isi_hijau = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    for i in range(2, ws.max_row + 1):
        keyakinan = ws.cell(row=i, column=col["Tingkat_Keyakinan"]).value
        warna = isi_hijau if keyakinan == "TINGGI" else (isi_kuning if keyakinan == "PERLU_CEK_MANUAL" else None)
        if warna:
            for c in range(1, ws.max_column + 1):
                ws.cell(row=i, column=c).fill = warna

    wb.save(FILE_OUTPUT)

    print("\n=== SELESAI ===")
    print(f"Lingkungan pasti (kata kunci, tanpa perlu API)  : {hitung['katakunci_pasti']}")
    print(f"Diproses lewat geocoding                         : {hitung['geocoding_dipakai']}")
    print(f"Kata kunci & geocoding SEPAKAT (keyakinan tinggi): {hitung['sepakat']}")
    print(f"Ambigu / gagal -> perlu cek manual (kontak ketua otomatis tersedia): {hitung['ambigu_atau_gagal']}")
    print(f"\nFile hasil: {FILE_OUTPUT}")
    print("Hijau = keyakinan tinggi (auto). Kuning = perlu cek manual (lihat kolom Kontak_Untuk_Konfirmasi).")


if __name__ == "__main__":
    main()
